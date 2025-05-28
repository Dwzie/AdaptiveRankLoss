import argparse
import numpy as np
import openpyxl
import torch
import torch.nn as nn
import joblib
import RERANKPRSM_upgrade_train
from RERANKPRSM_upgrade_train import hidden_size
from RERANKPRSM_upgrade_train import ResNeXt,ResNeXtBlock
from RERANKPRSM_upgrade_train import input_size,num_blocks,cardinality,num_classes
# 导入 precision_recall_curve, auc, 以及 roc_auc_score
from sklearn.metrics import precision_recall_curve, auc, roc_auc_score # <--- 添加 roc_auc_score
np.set_printoptions(threshold=np.inf)
gl=0

def input1(name):
    a,b,c= RERANKPRSM_upgrade_train.inputfile_np_withSeq(name)
    m1=joblib.load(filename="saved_model5/LR.pkl")
    m2 = joblib.load(filename="saved_model5/DecTree.pkl")
    m3 = joblib.load(filename="saved_model5/svm.pkl")
    m4 = joblib.load(filename="saved_model5/xgboost.pkl")
    LR_pred=m1.predict(a).reshape(-1,1)
    x1 = m2.predict_proba(a)[:, 1].reshape(-1, 1)
    x2 = m3.predict_proba(a)[:, 0].reshape(-1, 1)
    x3 = m4.predict_proba(a)[:, 1].reshape(-1, 1)
    X = np.concatenate((a, LR_pred, x1, x2, x3), axis=1)
    Y=torch.tensor(b).long()
    Z = torch.tensor(c).long()
    return torch.tensor(X),Y,Z

def predict(filename):
    model = ResNeXt(input_size=input_size, hidden_size=hidden_size, num_classes=num_classes, num_blocks=num_blocks,
                    cardinality=cardinality)
    x, y, z = input1(filename)
    model = joblib.load(filename="saved_model5/DP.pkl")
    x = x.to(torch.float32)
    score = model(x)

    # 转换为 numpy 数组
    score = score.detach().numpy().reshape(-1, 1)
    y = y.reshape(-1, 1)
    z = z.reshape(-1, 1)

    # 计算 Precision-Recall AUC
    y_true = y.flatten()
    y_scores = score.flatten()
    precision, recall, _ = precision_recall_curve(y_true, y_scores)
    pr_auc = auc(recall, precision)

    # --- Start: 添加 ROC AUC 计算 --- # <--- 添加
    roc_auc = roc_auc_score(y_true, y_scores)
    # --- End: 添加 ROC AUC 计算 --- # <--- 结束

    # 计算 FDR‐AUC
    sorted_indices = np.argsort(-y_scores)
    sorted_y = y_true[sorted_indices]
    cum_decoy = np.cumsum(sorted_y == 0)
    cum_target = np.cumsum(sorted_y == 1) + 1  # 避免除 0
    fdr_curve = cum_decoy / cum_target
    x_axis = np.linspace(0, 1, len(fdr_curve))
    fdr_auc = auc(x_axis, fdr_curve)


    score_label = np.concatenate((score, y, z), axis=1)
    toprank = sorted(score_label, key=lambda x: x[0], reverse=True)
    ppp = np.array(toprank)

    fdr = 0
    target = 1
    decoy = 0
    top = 0
    nixu = list()
    size = len(ppp)
    f = open('result_one3.txt', 'a')
    for i in ppp:
        if (i[1] == 1.0):
            target = target + 1
            fdr = decoy / (target)
            nixu.append(fdr)
        else:
            decoy = decoy + 1
            fdr = decoy / (target)
            nixu.append(fdr)
    nixu.reverse()
    for i in nixu:
        if i <= 0.01:
            top = size - 1 - nixu.index(i)
            break

    f.write(filename + '\n')
    f.write('重打分后有: ' + str(top) + '\n')

    A = list()
    B = list()
    for i in range(top):
        A.append(str(ppp[i][2]))

    e_fdr = RERANKPRSM_upgrade_train.getevalue_and_fdr(filename)
    cs = list()
    z = z.detach().numpy()
    for i in range(size):
        cs.append([e_fdr[i][0], e_fdr[i][1], z[i]])
    toprank2 = sorted(cs, key=lambda x: x[0], reverse=True)
    nixu = sorted(cs, key=lambda x: x[0], reverse=False)

    top = 0
    size = len(toprank2)
    for i in toprank2:
        if i[1] <= 0.01:
            top = size - 1 - toprank2.index(i)
            break
    for i in range(top):
        B.append(str(nixu[i][2]).strip("]").strip("["))

    f.write('之前有: ' + str(top) + '\n')

    C = 0
    C2 = 0
    cha = list()
    cha2 = list()
    chaprsm = list()
    chaprsm2 = list()
    for i in A:
        break_flag = False
        for j in B:
            if int(j) == int(float(i)):
                C = C + 1
                break_flag = True
                break
        if break_flag == True:
            continue
        cha.append(i)

    for i in B:
        break_flag = False
        for j in A:
            if int(i) == int(float(j)):
                C2 = C2 + 1
                break_flag = True
                break
        if break_flag == True:
            continue
        cha2.append(i)
    xx = x.numpy().tolist()

    for i in cha:  # PRSMREscore独有的
        chaprsm.append([str(int(xx[int(float(i))][0] * 34 + 0.1)), str(int(xx[int(float(i))][1] * 24 + 0.1)),
                        str(int(xx[int(float(i))][2] + 0.1)), str((e_fdr[int(float(i))][1])), i])
    for i in cha2:  # toppic独有的
        chaprsm2.append([str(int(xx[int(float(i))][0] * 34 + 0.1)), str(int(xx[int(float(i))][1] * 24 + 0.1)),
                         str(int(xx[int(float(i))][2] + 0.1)), str((e_fdr[int(float(i))][1])), i])

    chaprsm.append(["***", "***", "***", "***", "***"])
    chaprsm2.append(["***", "***", "***", "***", "***"])

    # 添加列名
    column_names = ["Matched Peaks", "Matched Fragments", "Normalized Matched Fragments", "FDR Value", "PRSM ID"]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(column_names)  # 添加列名
    for row_idx, row_data in enumerate(chaprsm, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx + 1, column=col_idx, value=cell_value)  # 从第二行开始写入数据
    output_filename = "PRSMREscore独有的.xlsx"
    workbook.save(output_filename)

    workbook2 = openpyxl.Workbook()
    sheet2 = workbook2.active
    sheet2.append(column_names)  # 添加列名
    for row_idx, row_data in enumerate(chaprsm2, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet2.cell(row=row_idx + 1, column=col_idx, value=cell_value)  # 从第二行开始写入数据
    output_filename = "toppic独有的.xlsx"
    workbook2.save(output_filename)

    f.write('共同有: ' + str(C) + '\n')
    f.write('Precision-Recall AUC: ' + str(pr_auc) + '\n')
    f.write('ROC AUC: ' + str(roc_auc) + '\n') # <--- 添加 ROC AUC 到输出文件
    f.write('FDR AUC: ' + str(fdr_auc) + '\n')
    f.write('\n')
    f.close()

    return "over!"

# NeuralNet 和 if __name__ == '__main__': 部分保持不变
class NeuralNet(nn.Module):
    def __init__(self, input_size, hidden_size, num_classes):
        super(NeuralNet, self).__init__()
        self.fc1 = nn.Linear(input_size, hidden_size)
        self.relu = nn.ReLU()
        self.fc2 = nn.Linear(hidden_size, num_classes)

    def forward(self, x):

        x=x.to(torch.float32)
        out = self.fc1(x)
        out = self.relu(out)
        out = self.fc2(out)
        return out


if __name__ == '__main__':
    # parser = argparse.ArgumentParser()
    # parser.add_argument('--arg1', type=str, help='参数1')
    # args = parser.parse_args()
    # #

    # predict("input1.6.2/FB_CB_1_ms2_toppic_prsm.xml")
    # predict("input1.6.2/FB_CB_2_ms2_toppic_prsm.xml")
    # predict("input1.6.2/FB_CB_4_ms2_toppic_prsm.xml")
    # predict("input1.6.2/FB_TeO_4_ms2_toppic_prsm.xml")
    # predict("input1.6.2/FB_TeO_5_ms2_toppic_prsm.xml")
    # predict("input1.6.2/FB_TeO_6_ms2_toppic_prsm.xml")
    #
    predict("input1.6.2/test/F1_1_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F2_1_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F1_2_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F2_2_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F3_1_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F4_1_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Yeast_3_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Yeast_5_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Yeast_6_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Yeast_7_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Yeast_8_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Yeast_9_ms2_toppic_prsm.xml")
    #
    predict("input1.6.2/test/081018_RVG262_PGAFF_RP4H_Neutros_GelFREE_8pF1_FAIMS_CV_-10_018_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/081018_RVG262_PGAFF_RP4H_Neutros_GelFREE_8pF1_FAIMS_CV_-20_005_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/081018_RVG262_PGAFF_RP4H_Neutros_GelFREE_8pF1_FAIMS_CV_-30_020_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/081018_RVG262_PGAFF_RP4H_Neutros_GelFREE_8pF1_FAIMS_CV_-50_019_ms2_toppic_prsm.xml")

    predict("input1.6.2/test/LCA_RM_20191005_Platelets_F1AB_01_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/LCA_RM_20191005_Platelets_F1AB_02_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/LCA_RM_20191005_Platelets_F2AB_01_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/LCA_RM_20191005_Platelets_F2AB_02_ms2_toppic_prsm.xml")
    #
    #
    predict("input1.6.2/test/ESVO_NSI_5939_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/ESVO_NSI_5942_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/ESVO_NSI_5951_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/ESVO_NSI_5954_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/ESVO_NSI_5957_ms2_toppic_prsm.xml")
    #
    #
    predict("input1.6.2/test/20181010_F1_Ag5_alban001_SA_TCx3_22_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/20181010_F1_Ag5_alban001_SA_TCx3_23_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/20181010_F1_Ag5_alban001_SA_TCx3_24_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/20181010_F1_Ag5_alban001_SA_TCx3_25_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/20181010_F1_Ag5_alban001_SA_TCx3_26_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/20181010_F1_Ag5_alban001_SA_TCx3_27_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/20181010_F1_Ag5_alban001_SA_TCx3_28_ms2_toppic_prsm.xml")


    predict("input1.6.2/test/C6_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F4_1_ms2_toppic_prsm_AT.xml")
    predict("input1.6.2/test/F4_2_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F5_1_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F6_1_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/F6_2_ms2_toppic_prsm.xml")

    predict("input1.6.2/test/Experiment_4_620_F1_01_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Experiment_4_620_F1_02_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Experiment_4_620_F1_03_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Experiment_4_620_F2_01_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Experiment_4_620_F2_02_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/Experiment_4_620_F2_03_ms2_toppic_prsm.xml")
    #
    predict("input1.6.2/test/FB_TeO_4_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/FB_TeO_5_ms2_toppic_prsm.xml")
    predict("input1.6.2/test/FB_TeO_6_ms2_toppic_prsm.xml")

